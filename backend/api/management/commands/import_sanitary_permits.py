import re
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand

from api.models import (
    PERMIT_STATUS_RENEWAL_DUE,
    RENEWAL_PAYMENT_UNPAID,
    RENEWAL_STAGE_LAPSED,
    SANITARY_STATUS_UPCOMING,
    SanitaryBusinessType,
    SanitaryEstablishment,
    SanitaryPermitRenewal,
)
from api.services.sanitation import generate_renewal_id

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


DEFAULT_FILES = [
    Path(r"C:\Users\This PC\Downloads\Sanitary Permit Large 2025.xlsx"),
    Path(r"C:\Users\This PC\Downloads\Sanitary Permit Sp 2025.xlsx"),
]

SHEET_CONFIG = {
    "FE": ("Food Establishment", "large", 5, 7, 8, 9),
    "Commercial Non Food": ("Commercial Non Food", "large", 4, 6, 7, 8),
    "WRS": ("Water Refilling Station", "large", 4, 6, 7, 8),
    "PP": ("Pool / Resort", "large", 4, 6, 7, 8),
    "IE": ("Industrial Establishment", "large", 4, 6, 7, 8),
    "AIE": ("Agricultural / Industrial Establishment", "large", 4, 6, 7, 8),
    "AFV": ("Ambulant Food Vendor", "sp", 3, 4, 5, 7),
    "PT": ("Public Transport", "sp", 3, 4, 5, 8),
    "FC": ("Food / Commercial", "sp", 3, 4, 5, 7),
    "FV-FB": ("Fishing Vessel / Boat", "sp", 3, 4, 5, 7),
    "CSW": ("Commercial Service Worker", "sp", 3, 4, 5, 6),
    "TIANGE": ("Tiange", "sp", 2, 3, 4, 5),
    "FOOD": ("Food Establishment", "sp", 3, 4, 5, 8),
}

PERMIT_PATTERN = re.compile(r"[A-Z]{1,5}-\d{3,4}-20\d{2}")


class Command(BaseCommand):
    help = "Import Mauban sanitary permit masterlist Excel files."

    def add_arguments(self, parser):
        parser.add_argument(
            "files",
            nargs="*",
            help="Optional Excel file paths. Defaults to the 2025 SP and Large files in Downloads.",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise RuntimeError("openpyxl is required to import sanitary permit files.")

        paths = [Path(path) for path in options["files"]] or DEFAULT_FILES
        total_created = 0
        total_updated = 0
        total_skipped = 0

        for path in paths:
            if not path.exists():
                self.stdout.write(self.style.WARNING(f"Skipped missing file: {path}"))
                continue

            created, updated, skipped = self.import_file(path)
            total_created += created
            total_updated += updated
            total_skipped += skipped
            self.stdout.write(
                self.style.SUCCESS(
                    f"{path.name}: {created} created, {updated} updated, {skipped} skipped"
                )
            )

        self.stdout.write(
            self.style.SUCCESS(
                f"Done. {total_created} establishments created, {total_updated} updated, {total_skipped} rows skipped."
            )
        )

    def import_file(self, path):
        workbook = load_workbook(path, data_only=True)
        created_count = 0
        updated_count = 0
        skipped_count = 0

        for sheet in workbook.worksheets:
            config = SHEET_CONFIG.get(sheet.title)
            if not config:
                continue

            business_type_name, permit_size, owner_col, barangay_col, issued_col, establishment_col = config
            business_type, _ = SanitaryBusinessType.objects.update_or_create(
                name=business_type_name,
                defaults={
                    "inspection_frequency": "monthly",
                    "description": f"Imported from {path.name} / {sheet.title}.",
                },
            )

            for row_number in range(1, sheet.max_row + 1):
                row = [sheet.cell(row_number, col).value for col in range(1, sheet.max_column + 1)]
                permit_number = find_permit_number(row)

                if not permit_number:
                    continue

                owner_name = clean_text(cell(row, owner_col))
                barangay = normalize_barangay(cell(row, barangay_col))
                establishment_name = clean_text(cell(row, establishment_col)) or owner_name
                issued_date = parse_date(cell(row, issued_col))

                if not owner_name or not establishment_name:
                    skipped_count += 1
                    continue

                establishment, created = SanitaryEstablishment.objects.update_or_create(
                    permit_number=permit_number,
                    defaults={
                        "business_name": establishment_name,
                        "owner_name": owner_name,
                        "business_type": business_type,
                        "permit_size": permit_size,
                        "barangay": barangay or "Unspecified",
                        "address": build_address(barangay),
                        "has_permit": True,
                        "permit_issued_date": issued_date,
                        "permit_expiry_date": date(2025, 12, 31),
                        "compliance_status": SANITARY_STATUS_UPCOMING,
                        "permit_status": PERMIT_STATUS_RENEWAL_DUE,
                        "remarks": f"Imported from {path.name} / {sheet.title}.",
                    },
                )

                sync_renewal(establishment, permit_number, issued_date)

                if created:
                    created_count += 1
                else:
                    updated_count += 1

        return created_count, updated_count, skipped_count


def sync_renewal(establishment, permit_number, issued_date):
    renewal = (
        SanitaryPermitRenewal.objects.filter(
            establishment=establishment,
            permit_number=permit_number,
        )
        .order_by("id")
        .first()
    )

    defaults = {
        "permit_type": "Sanitary Permit",
        "expiration_date": date(2025, 12, 31),
        "stage": RENEWAL_STAGE_LAPSED,
        "progress": 8,
        "renewal_fee": 500 if establishment.permit_size == "sp" else 1500,
        "payment_status": RENEWAL_PAYMENT_UNPAID,
        "submitted_requirements": ["Previous permit copy"],
        "inspection_status": "Imported 2025 permit; renewal follow-up required.",
        "remarks": f"Original 2025 permit issued {issued_date.isoformat() if issued_date else 'date not recorded'}.",
    }

    if renewal:
        for key, value in defaults.items():
            setattr(renewal, key, value)
        renewal.save()
        return renewal

    return SanitaryPermitRenewal.objects.create(
        renewal_id=generate_renewal_id(),
        establishment=establishment,
        permit_number=permit_number,
        **defaults,
    )


def find_permit_number(row):
    for value in row[:4]:
        text = clean_text(value)
        match = PERMIT_PATTERN.search(text)
        if match:
            return match.group(0)
    return ""


def cell(row, one_based_index):
    index = one_based_index - 1
    return row[index] if 0 <= index < len(row) else None


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\u2019", "'").strip().split())


def normalize_barangay(value):
    text = clean_text(value)
    text = re.sub(r"\bPOB\.?\b", "Poblacion", text, flags=re.IGNORECASE)
    return text.title()


def build_address(barangay):
    if not barangay:
        return "Mauban, Quezon"
    return f"Brgy. {normalize_barangay(barangay)}, Mauban, Quezon"


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None
