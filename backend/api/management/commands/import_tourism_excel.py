from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from api.models import Resort, ResortMonthlyArrival


MONTHS = {
    "JANUARY": 1,
    "FEBRUARY": 2,
    "MARCH": 3,
    "APRIL": 4,
    "MAY": 5,
    "JUNE": 6,
    "JULY": 7,
    "AUGUST": 8,
    "SEPTEMBER": 9,
    "OCTOBER": 10,
    "NOVEMBER": 11,
    "DECEMBER": 12,
}


class Command(BaseCommand):
    help = "Import tourism historical monthly resort arrivals from client Excel file."

    def add_arguments(self, parser):
        parser.add_argument(
            "excel_path",
            type=str,
            help="Path to MONITORING_FORM_2025.xlsx",
        )

        parser.add_argument(
            "--year",
            type=int,
            default=2025,
            help="Year of the monitoring data. Default is 2025.",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_path"])
        year = options["year"]

        if not excel_path.exists():
            self.stderr.write(self.style.ERROR(f"File not found: {excel_path}"))
            return

        workbook = load_workbook(excel_path, data_only=True)

        sheet_name = "Summary Resort Arrivals"

        if sheet_name not in workbook.sheetnames:
            self.stderr.write(
                self.style.ERROR(
                    f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"
                )
            )
            return

        sheet = workbook[sheet_name]

        imported_count = 0
        skipped_count = 0

        # Detect month columns from the first few rows.
        month_columns = {}

        for row in sheet.iter_rows(min_row=1, max_row=10):
            for cell in row:
                value = normalize_text(cell.value)

                if value in MONTHS:
                    month_columns[cell.column] = MONTHS[value]

        if not month_columns:
            self.stderr.write(
                self.style.ERROR(
                    "No month columns found. Please check the Summary Resort Arrivals sheet format."
                )
            )
            return

        # Try to find rows with resort names and monthly numeric values.
        for row in sheet.iter_rows(min_row=1):
            resort_name = None

            # Usually resort name is in the first few columns.
            for cell in row[:5]:
                text = normalize_text(cell.value)

                if text and text not in MONTHS and not text.isdigit():
                    # Skip obvious title/header words.
                    if text in ["RESORT", "TOTAL", "MONTH", "SUMMARY", "ARRIVALS"]:
                        continue

                    resort_name = str(cell.value).strip()
                    break

            if not resort_name:
                continue

            # Check if row has at least one month value.
            has_month_value = False

            for column_index, month_number in month_columns.items():
                cell_value = sheet.cell(row=row[0].row, column=column_index).value
                arrivals = parse_int(cell_value)

                if arrivals is not None:
                    has_month_value = True
                    break

            if not has_month_value:
                continue

            resort, _ = Resort.objects.get_or_create(
                resort_name=resort_name,
                defaults={
                    "resort_id": get_next_resort_id(),
                    "with_mayors_permit": False,
                    "type": "Resort",
                    "location": "Mauban, Quezon",
                    "short_description": "Imported from client monitoring form.",
                    "tourism_rating": 0,
                    "access": "N/A",
                    "itinerary_ids": [],
                    "image_key": "",
                    "monthly_arrivals": 0,
                    "latitude": 14.185,
                    "longitude": 121.731,
                },
            )

            for column_index, month_number in month_columns.items():
                cell_value = sheet.cell(row=row[0].row, column=column_index).value
                arrivals = parse_int(cell_value)

                if arrivals is None:
                    skipped_count += 1
                    continue

                ResortMonthlyArrival.objects.update_or_create(
                    resort=resort,
                    year=year,
                    month=month_number,
                    defaults={
                        "total_arrivals": arrivals,
                    },
                )

                imported_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete. Imported/updated: {imported_count}. Skipped blank/non-numeric cells: {skipped_count}."
            )
        )


def normalize_text(value):
    if value is None:
        return ""

    return str(value).strip().upper()


def parse_int(value):
    if value is None or value == "":
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    try:
        cleaned = str(value).replace(",", "").strip()

        if not cleaned:
            return None

        return int(float(cleaned))
    except (TypeError, ValueError):
        return None


def get_next_resort_id():
    latest = Resort.objects.order_by("-resort_id").first()

    if not latest:
        return 1

    return latest.resort_id + 1