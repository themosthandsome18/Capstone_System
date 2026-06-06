from datetime import datetime
import re

from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.validators import validate_email
from django.db import IntegrityError
from django.db.models import Max
from django.utils import timezone
from openpyxl import load_workbook

from ..models import (
    BOOKING_STATUS_CHOICES,
    BoatType,
    Country,
    Itinerary,
    Province,
    Region,
    Resort,
    TOURIST_RECORD_COUNT_FIELDS,
    TOURIST_RECORD_REQUIRED_FIELDS,
    TouristRecord,
    TravelMode,
    VisitPurpose,
    validate_tourist_record_values,
)
from ..seeders import ensure_initial_reference_data


COLUMNS = {
    "timestamp": 1,
    "email": 2,
    "consent": 3,
    "arrival_date": 4,
    "full_name": 5,
    "contact_number": 6,
    "resort": 7,
    "total_visitors": 8,
    "total_male": 9,
    "total_female": 10,
    "itinerary": 11,
    "maubanin_count": 12,
    "special_group_count": 13,
    "country": 14,
    "region": 15,
    "province": 16,
    "country_of_origin": 17,
    "foreigner_count": 18,
    "filipino_count": 19,
    "age_0_7": 20,
    "age_8_59": 21,
    "age_60_above": 22,
    "travel_mode": 23,
    "boat_type": 24,
    "boat_capacity_fare": 25,
    "parking_space": 26,
    "visit_purpose": 27,
}

RESORT_ALIASES = {
    "aguho playa beach resort": "Aguho Playa Beach Resort",
    "villa escaparde camping and beach resor": "Villa Escaparde Camping and Beach Resort",
    "villa escaparde camping and beach resort": "Villa Escaparde Camping and Beach Resort",
    "villa noe beach": "Villa Noe Beach",
    "villa noe beach resort": "Villa Noe Beach",
    "aquazul hotel and resort": "Aquazul Hotel and Resort",
    "dona choleng camping resort": "Dona Choleng Camping Resort",
    "doña choleng camping resort": "Dona Choleng Camping Resort",
    "tent place": "Tent Place",
    "villa cleofas": "Villa Cleofas Cagbalete",
    "villa cleofas cagbalete": "Villa Cleofas Cagbalete",
    "villa cleofas cagbalete island camping resort": "Villa Cleofas Cagbalete",
    "jovencios resort": "Jovencio's Resort",
    "jovencio's resort": "Jovencio's Resort",
    "mvt sto nino beach resort": "MVT Sto. Nino Beach Resort",
    "mvt sto. nino beach resort": "MVT Sto. Nino Beach Resort",
    "nenita del sol": "Nenita Del Sol",
    "tita pinay beach resort": "Tita Pinay Beach Resort",
    "orlan beach resort": "Orlan Beach Resort",
    "villa pilarosa beach resort": "Villa Pilarosa Beach Resort",
    "rio del sol beach resort": "Rio Del Sol Beach Resort",
    "nilandingan cove resort": "Nilandingan Cove Resort",
    "resident": "Residence",
    "residence": "Residence",
}

RELATED_FIELD_API_NAMES = {
    "country": "country_id",
    "region": "region_id",
    "province": "province_id",
    "itinerary": "itinerary_id",
    "resort": "resort_id",
    "travel_mode": "travel_mode_id",
    "boat_type": "boat_type_id",
    "visit_purpose": "visit_purpose_id",
}

BOOKING_STATUS_VALUES = {status for status, _label in BOOKING_STATUS_CHOICES}
BULK_BATCH_SIZE = 500
CREATE_RETRY_LIMIT = 5
ONLINE_BOOKING_ARRIVAL_YEARS = {2024, 2025, 2026}
TOURIST_RECORD_UPSERT_FIELDS = [
    "submitted_at",
    "email",
    "consent_confirmed",
    "full_name",
    "contact_number",
    "country",
    "region",
    "province",
    "country_of_origin",
    "foreigner_count",
    "filipino_count",
    "maubanin_count",
    "total_visitors",
    "total_male",
    "total_female",
    "special_group_count",
    "age_0_7",
    "age_8_59",
    "age_60_above",
    "arrival_date",
    "itinerary",
    "resort",
    "travel_mode",
    "boat_type",
    "boat_capacity_fare",
    "parking_space",
    "visit_purpose",
    "status",
    "updated_at",
]


def preview_online_booking_workbook(file_obj, status="pending", limit=0):
    return process_online_booking_workbook(
        file_obj,
        status=status,
        limit=limit,
        commit=False,
    )


def import_online_booking_workbook(file_obj, status="pending", limit=0):
    return process_online_booking_workbook(
        file_obj,
        status=status,
        limit=limit,
        commit=True,
    )


def process_online_booking_workbook(file_obj, status="pending", limit=0, commit=False):
    ensure_reference_data_ready()
    workbook = load_workbook(file_obj, data_only=True)
    sheet_name = "Form Responses 1"

    if sheet_name not in workbook.sheetnames:
        return {
            "valid_count": 0,
            "imported_count": 0,
            "updated_count": 0,
            "skipped_count": 0,
            "duplicate_count": 0,
            "new_resort_count": 0,
            "valid_samples": [],
            "error_samples": [
                {
                    "row": 0,
                    "message": f"Sheet '{sheet_name}' not found.",
                }
            ],
            "duplicate_samples": [],
            "new_resort_samples": [],
        }

    sheet = workbook[sheet_name]
    resolver = ReferenceResolver(commit=commit)
    existing_duplicate_keys, existing_survey_duplicate_keys = (
        build_existing_duplicate_keys()
    )
    workbook_duplicate_keys = set()
    valid_payloads = []
    valid_count = 0
    imported_count = 0
    updated_count = 0
    skipped_count = 0
    duplicate_count = 0
    valid_samples = []
    error_samples = []
    duplicate_samples = []
    max_row = sheet.max_row

    if limit:
        max_row = min(max_row, limit + 2)

    for row_number in range(3, max_row + 1):
        payload = build_payload(sheet, row_number, status, resolver)

        if not payload:
            skipped_count += 1
            continue

        duplicate_key = build_duplicate_key(payload, resolver)
        existing_survey_key = existing_survey_duplicate_keys.get(payload["survey_id"])
        is_duplicate_existing_record = (
            duplicate_key in existing_duplicate_keys
            and existing_survey_key != duplicate_key
        )
        is_duplicate_workbook_record = duplicate_key in workbook_duplicate_keys
        if is_duplicate_existing_record or is_duplicate_workbook_record:
            duplicate_count += 1
            skipped_count += 1
            if len(duplicate_samples) < 20:
                duplicate_samples.append(
                    {
                        "row": row_number,
                        "survey_id": payload["survey_id"],
                        "guest": payload["full_name"],
                        "contact": payload["contact_number"],
                        "arrival_date": payload["arrival_date"].isoformat(),
                        "resort": resolver.resort_name_by_id[payload["resort_id"]],
                        "message": "Possible duplicate booking with the same guest, contact, arrival date, and resort.",
                    }
                )
            continue

        workbook_duplicate_keys.add(duplicate_key)

        errors = validate_payload(payload)
        if errors:
            skipped_count += 1
            if len(error_samples) < 10:
                error_samples.append(
                    {
                        "row": row_number,
                        "guest": payload.get("full_name", ""),
                        "message": flatten_errors(errors),
                    }
                )
            continue

        valid_count += 1
        valid_payloads.append(payload)
        if len(valid_samples) < 10:
            valid_samples.append(
                {
                    "row": row_number,
                    "survey_id": payload["survey_id"],
                    "guest": payload["full_name"],
                    "arrival_date": payload["arrival_date"].isoformat(),
                    "resort": resolver.resort_name_by_id[payload["resort_id"]],
                    "total_visitors": payload["total_visitors"],
                }
            )

    if commit:
        imported_count, updated_count = persist_valid_records(valid_payloads)

    return {
        "valid_count": valid_count,
        "imported_count": imported_count,
        "updated_count": updated_count,
        "skipped_count": skipped_count,
        "duplicate_count": duplicate_count,
        "new_resort_count": len(resolver.new_resort_names),
        "valid_samples": valid_samples,
        "error_samples": error_samples,
        "duplicate_samples": duplicate_samples,
        "new_resort_samples": sorted(resolver.new_resort_names)[:20],
    }


def validate_payload(payload):
    values = {}

    for field in TOURIST_RECORD_REQUIRED_FIELDS:
        api_field = RELATED_FIELD_API_NAMES.get(field)
        values[field] = payload.get(api_field or field)

    for field in TOURIST_RECORD_COUNT_FIELDS:
        values[field] = payload.get(field, 0)

    errors = validate_tourist_record_values(values)
    email = payload.get("email")
    if email:
        try:
            validate_email(email)
        except DjangoValidationError:
            errors.setdefault("email", []).append("Enter a valid email address.")

    if payload.get("status") not in BOOKING_STATUS_VALUES:
        errors.setdefault("status", []).append("Invalid booking status.")

    arrival_date = payload.get("arrival_date")
    if (
        arrival_date
        and arrival_date.year not in ONLINE_BOOKING_ARRIVAL_YEARS
    ):
        errors.setdefault("arrival_date", []).append(
            "Must be a 2024, 2025, or 2026 online booking arrival date."
        )

    return {
        RELATED_FIELD_API_NAMES.get(field, field): messages
        for field, messages in errors.items()
    }


def persist_valid_records(valid_payloads):
    if not valid_payloads:
        return 0, 0

    survey_ids = [payload["survey_id"] for payload in valid_payloads]
    existing_survey_ids = set(
        TouristRecord.objects.filter(survey_id__in=survey_ids).values_list(
            "survey_id",
            flat=True,
        )
    )
    now = timezone.now()
    records = [
        TouristRecord(**payload, updated_at=now)
        for payload in valid_payloads
    ]

    TouristRecord.objects.bulk_create(
        records,
        batch_size=BULK_BATCH_SIZE,
        update_conflicts=True,
        update_fields=TOURIST_RECORD_UPSERT_FIELDS,
        unique_fields=["survey_id"],
    )

    updated_count = len(existing_survey_ids)
    imported_count = len(valid_payloads) - updated_count
    return imported_count, updated_count


def ensure_reference_data_ready():
    reference_models = [
        Country,
        Region,
        Province,
        Itinerary,
        TravelMode,
        BoatType,
        VisitPurpose,
        Resort,
    ]

    if all(model.objects.exists() for model in reference_models):
        return

    ensure_initial_reference_data()


def build_payload(sheet, row_number, status, resolver):
    full_name = clean_text(cell_value(sheet, row_number, "full_name"))
    arrival_date = parse_date(cell_value(sheet, row_number, "arrival_date"))
    contact_number = clean_text(cell_value(sheet, row_number, "contact_number"))

    if not full_name or not arrival_date or not contact_number:
        return None

    total_visitors = parse_int(cell_value(sheet, row_number, "total_visitors"))
    foreigner_count = parse_int(cell_value(sheet, row_number, "foreigner_count"))
    filipino_count = parse_int(cell_value(sheet, row_number, "filipino_count"))
    maubanin_count = parse_int(cell_value(sheet, row_number, "maubanin_count"))

    payload = {
        "survey_id": f"OBR-{arrival_date.year}-{row_number - 2:05d}",
        "submitted_at": parse_datetime(cell_value(sheet, row_number, "timestamp")),
        "email": clean_text(cell_value(sheet, row_number, "email")),
        "consent_confirmed": is_yes(cell_value(sheet, row_number, "consent")),
        "full_name": full_name,
        "contact_number": contact_number,
        "country_id": resolver.named(
            Country,
            normalize_country(cell_value(sheet, row_number, "country")),
            defaults={"type": "local"},
        ).id,
        "region_id": resolver.named(
            Region,
            normalize_region(cell_value(sheet, row_number, "region")),
        ).id,
        "province_id": resolver.named(
            Province,
            normalize_province(cell_value(sheet, row_number, "province")),
        ).id,
        "country_of_origin": clean_text(
            cell_value(sheet, row_number, "country_of_origin")
        ),
        "foreigner_count": foreigner_count,
        "filipino_count": filipino_count,
        "maubanin_count": maubanin_count,
        "total_visitors": total_visitors,
        "total_male": parse_int(cell_value(sheet, row_number, "total_male")),
        "total_female": parse_int(cell_value(sheet, row_number, "total_female")),
        "special_group_count": parse_int(
            cell_value(sheet, row_number, "special_group_count")
        ),
        "age_0_7": parse_int(cell_value(sheet, row_number, "age_0_7")),
        "age_8_59": parse_int(cell_value(sheet, row_number, "age_8_59")),
        "age_60_above": parse_int(cell_value(sheet, row_number, "age_60_above")),
        "arrival_date": arrival_date,
        "itinerary_id": resolver.named(
            Itinerary,
            normalize_itinerary(cell_value(sheet, row_number, "itinerary")),
        ).id,
        "resort_id": resolver.resort(
            clean_text(cell_value(sheet, row_number, "resort"))
        ).resort_id,
        "travel_mode_id": resolver.named(
            TravelMode,
            normalize_travel_mode(cell_value(sheet, row_number, "travel_mode")),
        ).id,
        "boat_type_id": resolver.named(
            BoatType,
            normalize_boat_type(cell_value(sheet, row_number, "boat_type")),
        ).id,
        "boat_capacity_fare": clean_text(
            cell_value(sheet, row_number, "boat_capacity_fare")
        ),
        "parking_space": clean_text(cell_value(sheet, row_number, "parking_space")),
        "visit_purpose_id": resolver.named(
            VisitPurpose,
            normalize_purpose(cell_value(sheet, row_number, "visit_purpose")),
        ).id,
        "status": status,
    }

    classification_total = foreigner_count + filipino_count + maubanin_count
    if not total_visitors and classification_total:
        payload["total_visitors"] = classification_total
        total_visitors = classification_total

    if classification_total != total_visitors:
        payload["filipino_count"] = max(
            total_visitors - foreigner_count - maubanin_count,
            0,
        )

    return payload


class ReferenceResolver:
    def __init__(self, commit=False):
        self.commit = commit
        self.named_cache = {}
        self.next_ids = {}
        self.new_resort_names = set()
        resorts = list(Resort.objects.all())
        self.resort_cache = {
            normalize_key(normalize_resort_name(resort.resort_name)): resort
            for resort in resorts
        }
        self.resort_name_by_id = {
            resort.resort_id: resort.resort_name for resort in resorts
        }
        self.next_resort_id = (
            Resort.objects.aggregate(max_id=Max("resort_id"))["max_id"] or 0
        ) + 1

    def named(self, model, name, defaults=None):
        defaults = defaults or {}
        name = clean_text(name) or "Unspecified"
        if model not in self.named_cache:
            self.named_cache[model] = {
                normalize_key(item.name): item for item in model.objects.all()
            }
        cache = self.named_cache[model]
        key = normalize_key(name)

        if key in cache:
            return cache[key]

        next_id = self.next_ids.get(model)
        if next_id is None:
            next_id = (model.objects.aggregate(max_id=Max("id"))["max_id"] or 0) + 1

        if self.commit:
            item = self.create_named_reference(model, next_id, name, defaults)
        else:
            item = model(id=next_id, name=name, **defaults)

        self.next_ids[model] = item.id + 1
        cache[key] = item
        return item

    def create_named_reference(self, model, next_id, name, defaults):
        for _attempt in range(CREATE_RETRY_LIMIT):
            try:
                return model.objects.create(id=next_id, name=name, **defaults)
            except IntegrityError:
                next_id = next_primary_key(model, "id")

        return model.objects.create(id=next_id, name=name, **defaults)

    def resort(self, resort_name):
        resort_name = normalize_resort_name(resort_name) or "Unspecified Resort"
        key = normalize_key(resort_name)

        if key in self.resort_cache:
            resort = self.resort_cache[key]
            self.resort_name_by_id[resort.resort_id] = resort.resort_name
            return resort

        resort_data = {
            "resort_id": self.next_resort_id,
            "resort_name": resort_name,
            "with_mayors_permit": False,
            "type": "Registered Resort",
            "location": "Mauban, Quezon",
            "short_description": "Imported from Tourism Office online booking responses.",
            "tourism_rating": 0,
            "access": "N/A",
            "itinerary_ids": [],
            "image_key": "",
            "monthly_arrivals": 0,
            "latitude": 14.185,
            "longitude": 121.731,
        }
        if self.commit:
            resort = self.create_resort(resort_data)
        else:
            resort = Resort(**resort_data)

        self.next_resort_id = resort.resort_id + 1
        self.resort_cache[key] = resort
        self.resort_name_by_id[resort.resort_id] = resort.resort_name
        self.new_resort_names.add(resort.resort_name)
        return resort

    def create_resort(self, resort_data):
        for _attempt in range(CREATE_RETRY_LIMIT):
            try:
                return Resort.objects.create(**resort_data)
            except IntegrityError:
                resort_data["resort_id"] = next_primary_key(
                    Resort,
                    "resort_id",
                )

        return Resort.objects.create(**resort_data)


def next_primary_key(model, field_name):
    return (model.objects.aggregate(max_id=Max(field_name))["max_id"] or 0) + 1


def build_existing_duplicate_keys():
    duplicate_keys = set()
    survey_duplicate_keys = {}
    records = TouristRecord.objects.select_related("resort").all()

    for record in records:
        key = (
            normalize_key(record.full_name),
            normalize_contact(record.contact_number),
            record.arrival_date.isoformat(),
            normalize_key(record.resort.resort_name if record.resort_id else ""),
        )
        duplicate_keys.add(key)
        survey_duplicate_keys[record.survey_id] = key

    return duplicate_keys, survey_duplicate_keys


def build_duplicate_key(payload, resolver):
    return (
        normalize_key(payload.get("full_name", "")),
        normalize_contact(payload.get("contact_number", "")),
        payload["arrival_date"].isoformat(),
        normalize_key(resolver.resort_name_by_id.get(payload.get("resort_id"), "")),
    )


def normalize_resort_name(value):
    text = clean_text(value)
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\b\d[\d\s/+()-]{5,}\b", "", text)
    text = clean_text(text).strip(" ,-")
    key = normalize_key(text).replace(".", "")
    return RESORT_ALIASES.get(key, text)


def flatten_errors(errors):
    parts = []
    for field, messages in errors.items():
        if isinstance(messages, (list, tuple)):
            text = " ".join(str(message) for message in messages)
        else:
            text = str(messages)
        parts.append(f"{field}: {text}")
    return " ".join(parts)


def cell_value(sheet, row_number, key):
    return sheet.cell(row=row_number, column=COLUMNS[key]).value


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_key(value):
    return clean_text(value).lower()


def normalize_contact(value):
    text = clean_text(value)
    return re.sub(r"\D+", "", text)


def normalize_country(value):
    text = clean_text(value)
    if text.lower().startswith("phil"):
        return "Philippines"
    if not text:
        return "Philippines"
    return text.title()


def normalize_region(value):
    text = clean_text(value)
    known = {
        "calabarzon region": "CALABARZON Region",
        "ncr - national capital region": "NCR - National Capital Region",
        "region iii - central luzon": "Region III - Central Luzon",
        "region v - bicol region": "Region V - Bicol Region",
        "mimaropa region": "MIMAROPA Region",
        "car - cordillera administrative region": "CAR - Cordillera Administrative Region",
    }
    return known.get(normalize_key(text), text or "CALABARZON Region")


def normalize_province(value):
    text = clean_text(value).replace("Province", "").replace("province", "")
    normalized = normalize_key(text)
    if normalized in {"ncr", "metro manila", "metro manila."}:
        return "Metro Manila"
    return clean_text(text).title() or "Quezon"


def normalize_itinerary(value):
    text = clean_text(value)
    key = normalize_key(text).replace(" ", "")
    if key in {"2nights", "2night"}:
        return "2 Nights"
    if key in {"3nights", "3night"}:
        return "3 Nights"
    if key in {"4nights", "4night"}:
        return "4 Nights"
    if "5" in key and "night" in key:
        return "5 Nights and above"
    if "day" in key:
        return "Day Tour"
    if "overnight" in key or not key:
        return "Overnight"
    return text.title()


def normalize_travel_mode(value):
    text = clean_text(value)
    if normalize_key(text) == "public utility":
        return "Public Utility Vehicle"
    return text or "Private Vehicle"


def normalize_boat_type(value):
    text = clean_text(value)
    key = normalize_key(text)
    if key.startswith("private boat"):
        return "Private Boat (Rates depend on the capacity)"
    if key.startswith("public boat"):
        return "Public Boat (P100/ride/head) Sabang Port Only"
    if key.startswith("boat provided"):
        return "Boat Provided by Resort (As confirmed by both guests and resort)"
    return text or "Public Boat (P100/ride/head) Sabang Port Only"


def normalize_purpose(value):
    text = clean_text(value)
    key = normalize_key(text)
    if key == "team building":
        return "Team Building"
    if not text:
        return "Leisure"
    return text.title()


def parse_int(value):
    if value in (None, ""):
        return 0
    text = clean_text(value)
    if not text or text.lower() in {"none", "n/a", "na", "."}:
        return 0
    try:
        return max(int(float(text.replace(",", ""))), 0)
    except ValueError:
        return 0


def parse_date(value):
    if isinstance(value, datetime):
        return normalize_arrival_date_year(value.date())
    text = clean_text(value)
    if not text:
        return None
    for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return normalize_arrival_date_year(
                datetime.strptime(text, date_format).date()
            )
        except ValueError:
            continue
    return None


def normalize_arrival_date_year(value):
    year = value.year

    if 1 <= year < 1000:
        year = 2000 + (year % 100)
    elif 2900 <= year <= 2999:
        year -= 900

    if year != value.year:
        return value.replace(year=year)

    return value


def parse_datetime(value):
    if isinstance(value, datetime):
        if timezone.is_naive(value):
            return timezone.make_aware(value)
        return value
    return None


def is_yes(value):
    return normalize_key(value) in {"yes", "y", "true", "1"}
